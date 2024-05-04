import logging

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from fake_useragent import UserAgent


class WebDriverHandler:
    def __init__(self, web, chromedriver_path):
        self.web = web
        self.chromedriver_path = chromedriver_path
        self.driver = None
        self.finished = None

    def __enter__(self):
        service = Service(executable_path=self.chromedriver_path)
        ua = UserAgent()
        a = ua.random
        user_agent = ua.random
        print(user_agent)
        options = webdriver.ChromeOptions()
        options.add_argument(f'user-agent={user_agent}')

        options.add_argument('--disable-blink-features=AutomationControlled')

        options.add_argument('--headless')
        options.add_argument("--window-size=1920,1080")
        self.driver = webdriver.Chrome(service=service, options=options)
        return self.driver

    def __exit__(self, type, value, traceback):
        if self.finished is True:
            self.driver.quit()


def start_scrapping(web, chromedriver_path):
    try:
        if not (web and chromedriver_path):
            raise ValueError('Web and chromedriver_path must be provided')

        with WebDriverHandler(web, chromedriver_path) as driver:
            driver.finished = False
            logging.info(f"Opening webpage: {web}")
            driver.get(web)

            products_category_list_elements = driver.find_elements(by='xpath', value='//div//ul[@class="inside"]')
            products_category_list_elements_ref = driver.find_elements(by='xpath', value='//div//ul[@class="inside"]//li//a')
            products_category_list_elements_links = [elem.get_attribute('href') for elem in products_category_list_elements_ref]
            products_category_list_text = [element.text for element in products_category_list_elements]
            products_header_elements = driver.find_elements(by='xpath', value='//div[contains(@class,"subTitle__header")]//h2')
            products_header_text = [element.text for element in products_header_elements]

            products_dictionary = {}
            for category, string in zip(products_header_text, products_category_list_text):
                cleaned_values = [value.strip() for value in string.split('\n')]
                cleaned_values_with_url = [{'name': name, 'url': products_category_list_elements_links[index]} for index, name in
                                           enumerate(cleaned_values)]
                products_dictionary[category] = cleaned_values_with_url

            cpu = products_dictionary[products_header_text[0]][0]['url']
            driver.get(cpu)
            # # Create a new Workbook object
            # wb = Workbook()
            #
            # # Select the active worksheet
            # ws = wb.active
            #
            # # Write headers to the worksheet
            # ws['A1'] = 'Name'
            # ws['B1'] = 'Author'
            # ws['C1'] = 'Runtime'
            #
            # row = 2

        # products = driver.find_elements(by='xpath', value='//li[contains(@class,"productListItem")]')
        # for product in products:
        #     name = product.find_element(by='xpath', value='.//h3[contains(@class,"bc-heading")]').text
        #     author = product.find_element(by='xpath', value='.//li[contains(@class,"authorLabel")]//a').text
        #     runtime = product.find_element(by='xpath', value='.//li[contains(@class,"runtimeLabel")]').text
        #
        #     ws[f'A{row}'] = name
        #     ws[f'B{row}'] = author
        #     ws[f'C{row}'] = runtime.split(':')[1].strip()
        #     row += 1
        #
        # wb.save('scraped_data.xlsx')
        driver.finished = False
    except Exception as e:
        logging.error(f"An error occurred: {e}")
